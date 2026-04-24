from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_XLSX = ROOT / "Raman_krov_SSZ-zdorovye.xlsx"
OUTPUT_DIR = ROOT / "patient_csv_exports"

SHEET_TO_GROUP = {
    "health": "healthy",
    "heart disease": "disease",
}

SELECTED_PATIENTS = {
    "health": ["healthy1", "healthy2", "healthy3", "healthy4", "healthy5"],
    "heart disease": [
        "heart_patient1",
        "heart_patient2",
        "heart_patient3",
        "heart_patient4",
        "heart_patient5",
    ],
}


def export_sheet(sheet_name: str, group_name: str, wb) -> list[str]:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = rows[0]
    data_rows = rows[1:]
    wavenumber = [float(row[0]) for row in data_rows if row[0] is not None]
    selected_names = set(SELECTED_PATIENTS[sheet_name])

    group_dir = OUTPUT_DIR / group_name
    group_dir.mkdir(parents=True, exist_ok=True)

    exported_names: list[str] = []
    for col_idx, patient_name in enumerate(header[1:], start=1):
        if patient_name is None:
            continue
        patient_name = str(patient_name).strip()
        if patient_name not in selected_names:
            continue

        intensities = [row[col_idx] for row in data_rows]
        if any(value is None for value in intensities):
            continue

        patient_slug = patient_name.replace(" ", "_")
        out_path = group_dir / f"{patient_slug}.csv"

        with out_path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(["wavenumber", "intensity"])
            for wn, intensity in zip(wavenumber, intensities):
                writer.writerow([f"{float(wn):.12f}", f"{float(intensity):.12f}"])

        exported_names.append(patient_name)

    return exported_names


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(SOURCE_XLSX, read_only=True, data_only=True)

    summary_rows: list[tuple[str, str, str]] = []
    for sheet_name, group_name in SHEET_TO_GROUP.items():
        exported_names = export_sheet(sheet_name, group_name, wb)
        for patient_name in exported_names:
            summary_rows.append(
                (group_name, patient_name, f"{group_name}/{patient_name.replace(' ', '_')}.csv")
            )

    manifest_path = OUTPUT_DIR / "manifest.csv"
    with manifest_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["group", "patient_name", "relative_path"])
        writer.writerows(summary_rows)

    print(f"Exported {len(summary_rows)} spectra to {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
